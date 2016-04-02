import openpyxl
import numpy as np
import seaborn as sns
import matplotlib.pyplot as plt

sns.set(style="white", context="talk")
rs = np.random.RandomState(7)

wb = openpyxl.load_workbook('finance_and_giving.xlsx')
sheet1 = wb.get_sheet_by_name("12.1.1") #UC_Total_03to15

y2005 = []
y2010 = []
y2014 = []
for c in range(0,2):
    y2005.append([])
    y2010.append([])
    y2014.append([])
    for r in range(3, 10):
        if c == 0:
            y2005[0].append(int(sheet1.cell(row = r, column = 5).value))
            y2010[0].append(int(sheet1.cell(row = r, column = 10).value))
            y2014[0].append(int(sheet1.cell(row = r, column = 14).value))
        elif c ==1:
            y2005[1].append(sheet1.cell(row = r, column = 2).value)
            y2010[1].append(sheet1.cell(row = r, column = 2).value)
            y2014[1].append(sheet1.cell(row = r, column = 2).value)

fig1 = sns.barplot(y2005[1], y2005[0], palette="BuGn_d")
plt.xlabel('Source of Income', fontsize = 20)
plt.ylabel('Amount Income in Dollars', fontsize = 20)
plt.title('Source of UC Income from 2013- 2014', fontsize = 24)

ax = plt.subplot(111)
ax.spines["top"].set_visible(False)
ax.spines["bottom"].set_visible(False)
ax.spines["right"].set_visible(False)
ax.spines["left"].set_visible(False)

f , (ax1, ax2, ax3) = plt.subplots(3, 1, figsize=(8, 6), sharex=True)

sns.barplot(y2005[1], y2005[0], palette="BuGn_d", ax=ax1)
ax1.set_ylabel('2004-2005', fontsize = 16)

sns.barplot(y2010[1], y2010[0], palette="RdBu_r", ax=ax2)
ax2.set_ylabel('2009-2010', fontsize = 16)

sns.barplot(y2014[1], y2014[0], palette="Set3", ax=ax3)
#ax3.set_ylabel("2014")
ax3.set_ylabel('2013-2014', fontsize = 16)

sns.despine(bottom=True)
plt.setp(f.axes, yticks=[])
plt.xticks(fontsize=10)
plt.tight_layout(h_pad=3)
plt.text
plt.show()
